"""
Excel中英互译脚本
功能：用户可选择要翻译的列和翻译结果填入的列，自动识别中文或英文后互译（中文→英文，英文→中文）
作者：AI助手
"""

import openpyxl  # 用于读写Excel文件
import requests  # 用于发送HTTP请求调用API
import hashlib  # 用于生成MD5或SHA256签名
import time  # 用于生成时间戳
import random  # 用于生成随机数（salt）
import re  # 用于正则表达式，判断是否为中文
import string  # 用于列号字母转数字

# ==================== 配置区域 ====================
# 有道翻译API配置
YOUDAO_APP_KEY = '1ffaa7e3a571d047'  # 替换为你的有道AppKey
YOUDAO_APP_SECRET = 'gIt1cpx6wToGX2UFIYDRWplVRP2k3umJ'  # 替换为你的有道AppSecret
YOUDAO_API_URL = 'https://openapi.youdao.com/api'

# DeepL翻译API配置
DEEPL_API_KEY = '28df4eab-f4e0-47cb-92a6-784b364b6e9e:fx'  # 替换为你的DeepL API密钥
DEEPL_API_URL = 'https://api-free.deepl.com/v2/translate'  # DeepL免费版API地址
# 如果使用DeepL Pro（付费版），使用：'https://api.deepl.com/v2/translate'

# Excel文件路径
EXCEL_FILE = '中英互译测试.xlsx'  # 可以修改为你需要翻译的Excel文件名

# 翻译延时设置（秒）
TRANSLATE_DELAY = 1.0  # 每次翻译之间的延时（秒），固定1秒
RETRY_DELAY = 1.0  # 遇到频率限制错误时的重试延时（秒），固定1秒
MAX_RETRIES = 3  # 遇到频率限制错误时的最大重试次数
# ================================================


def detect_language(text):
    """
    自动检测文本是中文还是英文
    
    参数：
        text: 要检测的文本
    
    返回：
        'zh' 表示中文，'en' 表示英文，'unknown' 表示无法判断
    """
    # 使用正则表达式匹配中文字符（Unicode范围：\u4e00-\u9fff）
    chinese_pattern = re.compile(r'[\u4e00-\u9fff]')
    
    # 统计文本中的中文字符数量
    chinese_chars = len(chinese_pattern.findall(text))
    
    # 统计文本中英文字母和数字的数量
    english_chars = len(re.findall(r'[a-zA-Z0-9]', text))
    
    # 如果中文字符数量 > 0，且中文字符数量 > 英文字符数量的一半，判定为中文
    # 这样可以避免一些混合文本被误判
    if chinese_chars > 0 and chinese_chars >= english_chars * 0.3:
        return 'zh'  # 中文
    elif english_chars > 0:
        return 'en'  # 英文
    else:
        return 'unknown'  # 无法判断（可能是数字或其他字符）


def get_error_message(error_code):
    """
    根据错误代码返回友好的错误提示信息
    
    参数：
        error_code: 有道翻译API返回的错误代码（字符串格式）
    
    返回：
        错误提示信息
    """
    error_codes = {
        '101': '缺少必填参数，请检查请求参数',
        '102': '不支持的语言类型',
        '103': '翻译文本过长（超过5000字符）',
        '104': '不支持的API类型',
        '105': '不支持的签名类型',
        '106': '不支持的响应类型',
        '107': '不支持的传输加密类型',
        '108': 'appKey无效，请检查API密钥',
        '109': '签名校验失败，请检查API密钥',
        '110': '无相关服务的有效实例',
        '111': '开发者账号无效',
        '201': '请求被拒绝，可能是API密钥权限不足',
        '202': '请求频率超限，请稍后再试',
        '301': '辞典查询失败',
        '302': '翻译查询失败',
        '303': '服务端的其它异常',
        '401': '账户余额不足',
        '411': '文本过长错误或请求频率受限（如果文本>2000字符则为文本过长）',
        '412': '长请求过于频繁，请稍后再试'
    }
    return error_codes.get(str(error_code), f'未知错误（错误代码：{error_code}）')


def translate_text_youdao(text, from_lang='zh-CHS', to_lang='en', retry_count=0):
    """
    调用有道翻译API翻译文本（带重试机制）
    
    参数：
        text: 要翻译的文本
        from_lang: 源语言，默认是中文（有道格式：zh-CHS, en）
        to_lang: 目标语言，默认是英文（有道格式：zh-CHS, en）
        retry_count: 当前重试次数（内部使用）
    
    返回：
        翻译后的文本，如果失败返回None
    """
    try:
        # 检查文本长度（有道翻译API实际限制，文本过长会导致411错误）
        text_length = len(text)
        if text_length > 5000:
            print(f"  ❌ 文本过长错误：文本长度 {text_length} 字符，超过5000字符限制，请缩短文本")
            return None
        elif text_length > 2000:
            # 文本过长，会直接返回错误，不发送请求
            print(f"  ❌ 文本过长错误：文本长度 {text_length} 字符，超过建议长度2000字符")
            print(f"     💡 建议：请将文本缩短至2000字符以内，或分段处理")
            return None
        
        # 检查文本是否为空
        if not text or not text.strip():
            print("❌ 文本为空，跳过翻译")
            return None
        
        # 生成随机数作为salt（盐值），用于加密签名
        salt = str(random.randint(1, 65536))
        
        # 获取当前时间戳（秒级）
        curtime = str(int(time.time()))
        
        # 计算签名：如果文本长度超过200字符，需要截取前10个字符和最后10个字符参与签名
        # 这是有道翻译API v3的要求，避免签名字符串过长
        input_text = text
        if len(input_text) > 200:
            input_text = input_text[:10] + str(len(text)) + input_text[-10:]
        
        # 拼接签名字符串：appKey + 原文（或截取后的） + salt + 时间戳 + appSecret
        sign_str = YOUDAO_APP_KEY + input_text + salt + curtime + YOUDAO_APP_SECRET
        
        # 使用SHA256算法对签名字符串进行加密，得到签名
        sign = hashlib.sha256(sign_str.encode('utf-8')).hexdigest()
        
        # 准备API请求的参数
        data = {
            'q': text,  # 要翻译的文本（完整文本）
            'from': from_lang,  # 源语言
            'to': to_lang,  # 目标语言
            'appKey': YOUDAO_APP_KEY,  # 应用ID
            'salt': salt,  # 随机数
            'sign': sign,  # 签名
            'signType': 'v3',  # 签名类型，v3表示使用SHA256
            'curtime': curtime  # 时间戳
        }
        
        # 发送POST请求到有道翻译API（推荐使用POST，避免URL长度限制）
        # 有道翻译API v3支持POST请求，使用POST可以避免URL长度限制问题
        response = requests.post(YOUDAO_API_URL, data=data, timeout=10)
        
        # 检查HTTP状态码
        if response.status_code != 200:
            print(f"❌ HTTP请求失败，状态码：{response.status_code}")
            return None
        
        # 将返回的JSON格式数据转换为Python字典
        result = response.json()
        
        # 检查返回结果中是否有错误代码
        error_code = result.get('errorCode')
        if error_code != '0' and error_code != 0:
            # 如果有错误，打印详细的错误信息
            error_msg = result.get('msg', '')
            friendly_msg = get_error_message(error_code)
            print(f"  ❌ 翻译失败：{friendly_msg}")
            if error_msg:
                print(f"     详细错误：{error_msg}")
            
            # 对于411错误，检查是否因为文本过长
            if str(error_code) == '411':
                text_length = len(text)
                print(f"     当前文本长度：{text_length} 字符")
                # 如果文本超过2000字符，判定为文本过长错误
                if text_length > 2000:
                    print(f"     ❌ 文本过长错误：文本长度 {text_length} 字符，超过2000字符限制")
                    print(f"     💡 建议：请将文本缩短至2000字符以内，或分段处理")
                    return None  # 文本过长时，不重试，直接返回
                else:
                    # 文本长度正常，可能是频率限制，使用固定1秒延时重试
                    if retry_count < MAX_RETRIES:
                        print(f"     ⏳ 可能是频率限制，等待 {RETRY_DELAY} 秒后自动重试（第 {retry_count + 1}/{MAX_RETRIES} 次）...")
                        time.sleep(RETRY_DELAY)  # 固定1秒延时
                        return translate_text_youdao(text, from_lang, to_lang, retry_count + 1)
                    else:
                        print(f"     ❌ 已达到最大重试次数（{MAX_RETRIES}次）")
                        print(f"     💡 建议：等待几分钟后重新运行程序")
            
            # 对于其他频率限制错误（202、412），先检查文本长度
            elif str(error_code) in ['202', '412']:
                text_length = len(text)
                print(f"     当前文本长度：{text_length} 字符")
                
                # 如果文本很长（>2000字符），可能是文本过长导致的错误，不重试
                if text_length > 2000:
                    print(f"     ❌ 文本过长错误：虽然返回{error_code}错误，但文本长度 {text_length} 字符超过2000字符限制")
                    print(f"     💡 建议：请将文本缩短至2000字符以内，或分段处理")
                    return None  # 文本过长时，不重试，直接返回
                
                # 文本长度正常，进行频率限制重试（固定1秒延时）
                if retry_count < MAX_RETRIES:
                    print(f"     ⏳ 频率限制错误，等待 {RETRY_DELAY} 秒后自动重试（第 {retry_count + 1}/{MAX_RETRIES} 次）...")
                    time.sleep(RETRY_DELAY)  # 固定1秒延时
                    # 递归重试
                    return translate_text_youdao(text, from_lang, to_lang, retry_count + 1)
                else:
                    print(f"     ❌ 已达到最大重试次数（{MAX_RETRIES}次）")
                    print(f"     💡 建议：等待几分钟后重新运行程序")
            
            return None
        
        # 提取翻译结果（返回的是一个列表，取第一个元素）
        if 'translation' in result and len(result['translation']) > 0:
            return result['translation'][0]
        else:
            print(f"  ❌ 翻译结果格式异常：{result}")
            return None
            
    except requests.exceptions.Timeout:
        print(f"  ❌ 翻译请求超时，请检查网络连接")
        return None
    except requests.exceptions.RequestException as e:
        print(f"  ❌ 网络请求异常：{str(e)}")
        return None
    except Exception as e:
        # 如果出现其他异常（比如JSON解析错误），打印错误信息
        print(f"  ❌ 翻译过程中出现异常：{str(e)}")
        return None


def translate_text_deepl(text, from_lang='ZH', to_lang='EN', retry_count=0):
    """
    调用DeepL翻译API翻译文本（带重试机制）
    
    参数：
        text: 要翻译的文本
        from_lang: 源语言，默认是中文（DeepL格式：ZH, EN，或使用auto自动检测）
        to_lang: 目标语言，默认是英文（DeepL格式：ZH, EN）
        retry_count: 当前重试次数（内部使用）
    
    返回：
        翻译后的文本，如果失败返回None
    """
    try:
        # 检查文本长度（DeepL免费版限制单次翻译文本不超过5000字符）
        text_length = len(text)
        if text_length > 5000:
            print(f"  ❌ 文本过长错误：文本长度 {text_length} 字符，超过5000字符限制，请缩短文本")
            return None
        elif text_length > 2000:
            # 文本过长，会直接返回错误，不发送请求
            print(f"  ❌ 文本过长错误：文本长度 {text_length} 字符，超过建议长度2000字符")
            print(f"     💡 建议：请将文本缩短至2000字符以内，或分段处理")
            return None
        
        # 检查文本是否为空
        if not text or not text.strip():
            print("❌ 文本为空，跳过翻译")
            return None
        
        # 准备API请求的参数
        # DeepL API：source_lang可以使用'auto'自动检测，也可以指定语言
        data = {
            'auth_key': DEEPL_API_KEY,  # DeepL API密钥
            'text': text,  # 要翻译的文本
            'target_lang': to_lang,  # 目标语言（必需）
        }
        
        # DeepL支持自动检测源语言，如果from_lang不是'auto'，则指定源语言
        # 但为了更好的准确性，我们使用自动检测（'auto'）
        if from_lang and from_lang.upper() != 'AUTO':
            data['source_lang'] = from_lang  # 如果明确指定了源语言，则使用指定值
        # 如果from_lang是'auto'，则不添加source_lang参数，让DeepL自动检测
        
        # 发送POST请求到DeepL翻译API
        response = requests.post(DEEPL_API_URL, data=data, timeout=10)
        
        # 检查HTTP状态码
        if response.status_code != 200:
            print(f"  ❌ HTTP请求失败，状态码：{response.status_code}")
            if response.status_code == 403:
                print(f"     💡 提示：可能是API密钥无效或权限不足")
            elif response.status_code == 456:
                print(f"     💡 提示：本月字符配额已用完")
            return None
        
        # 将返回的JSON格式数据转换为Python字典
        result = response.json()
        
        # 检查返回结果中是否有错误
        if 'message' in result:
            error_msg = result.get('message', '')
            print(f"  ❌ DeepL翻译失败：{error_msg}")
            
            # 如果是配额或频率限制错误，尝试重试
            if response.status_code == 429 or 'quota' in error_msg.lower() or 'limit' in error_msg.lower():
                if retry_count < MAX_RETRIES:
                    print(f"     ⏳ 频率限制错误，等待 {RETRY_DELAY} 秒后自动重试（第 {retry_count + 1}/{MAX_RETRIES} 次）...")
                    time.sleep(RETRY_DELAY)  # 固定1秒延时
                    return translate_text_deepl(text, from_lang, to_lang, retry_count + 1)
                else:
                    print(f"     ❌ 已达到最大重试次数（{MAX_RETRIES}次）")
                    print(f"     💡 建议：等待几分钟后重新运行程序")
            
            return None
        
        # 提取翻译结果
        if 'translations' in result and len(result['translations']) > 0:
            return result['translations'][0].get('text', None)
        else:
            print(f"  ❌ DeepL翻译结果格式异常：{result}")
            return None
            
    except requests.exceptions.Timeout:
        print(f"  ❌ 翻译请求超时，请检查网络连接")
        return None
    except requests.exceptions.RequestException as e:
        print(f"  ❌ 网络请求异常：{str(e)}")
        return None
    except Exception as e:
        # 如果出现其他异常（比如JSON解析错误），打印错误信息
        print(f"  ❌ 翻译过程中出现异常：{str(e)}")
        return None


def convert_lang_code_to_youdao(lang_code):
    """
    将语言代码转换为有道翻译API格式
    
    参数：
        lang_code: 语言代码（'zh' 或 'en'）
    
    返回：
        有道翻译API格式的语言代码
    """
    if lang_code == 'zh':
        return 'zh-CHS'  # 有道使用zh-CHS表示中文
    elif lang_code == 'en':
        return 'en'  # 有道使用en表示英文
    else:
        return lang_code


def convert_lang_code_to_deepl(lang_code):
    """
    将语言代码转换为DeepL翻译API格式
    
    参数：
        lang_code: 语言代码（'zh' 或 'en'）
    
    返回：
        DeepL翻译API格式的语言代码
    """
    if lang_code == 'zh':
        return 'ZH'  # DeepL使用ZH表示中文
    elif lang_code == 'en':
        return 'EN'  # DeepL使用EN表示英文
    else:
        return lang_code.upper()  # DeepL使用大写


def translate_text(text, from_lang_code, to_lang_code, service='youdao'):
    """
    统一的翻译接口，根据选择的服务调用相应的翻译函数
    
    参数：
        text: 要翻译的文本
        from_lang_code: 源语言代码（'zh' 或 'en'），用于有道翻译
        to_lang_code: 目标语言代码（'zh' 或 'en'）
        service: 翻译服务（'youdao' 或 'deepl'）
    
    返回：
        翻译后的文本，如果失败返回None
    """
    if service == 'youdao':
        # 转换为有道翻译API的语言代码格式
        from_lang = convert_lang_code_to_youdao(from_lang_code)
        to_lang = convert_lang_code_to_youdao(to_lang_code)
        return translate_text_youdao(text, from_lang, to_lang)
    elif service == 'deepl':
        # DeepL可以自动检测源语言，所以不传source_lang参数（或者传'auto'）
        # 只需要目标语言
        to_lang = convert_lang_code_to_deepl(to_lang_code)
        # 使用'auto'让DeepL自动检测源语言，这样更智能
        return translate_text_deepl(text, 'auto', to_lang)
    else:
        print(f"  ❌ 不支持的翻译服务：{service}")
        return None


def column_letter_to_number(column_input):
    """
    将列号转换为数字（支持字母格式如A、B、C，也支持数字格式如1、2、3）
    
    参数：
        column_input: 用户输入的列号（可以是'A'、'B'或'1'、'2'等）
    
    返回：
        列的数字编号（1表示A列，2表示B列，以此类推）
    """
    column_input = str(column_input).strip().upper()  # 转换为大写并去除空格
    
    # 如果输入的是字母（如A、B、C）
    if column_input.isalpha():
        # 将字母转换为数字：A=1, B=2, ..., Z=26, AA=27, AB=28, ...
        result = 0
        for char in column_input:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result
    # 如果输入的是数字（如1、2、3）
    elif column_input.isdigit():
        return int(column_input)
    else:
        return None  # 无效输入


def number_to_column_letter(column_num):
    """
    将列的数字编号转换为字母格式（1→A, 2→B, 等等）
    
    参数：
        column_num: 列的数字编号
    
    返回：
        列的字母表示（如'A'、'B'、'AA'等）
    """
    result = ""
    while column_num > 0:
        column_num -= 1  # 转换为0-based索引
        result = chr(65 + (column_num % 26)) + result  # 65是'A'的ASCII码
        column_num //= 26
    return result


def show_excel_preview(sheet, max_cols=5):
    """
    显示Excel文件的前几列和行的预览，帮助用户了解文件结构
    
    参数：
        sheet: Excel工作表对象
        max_cols: 最多显示多少列（默认5列）
    """
    max_row = min(sheet.max_row, 5)  # 最多显示5行
    max_col = min(sheet.max_column, max_cols)  # 最多显示指定列数
    
    print("\n📊 Excel文件预览（前5行）：")
    print("-" * 60)
    
    # 显示列标题
    header = "行号"
    for col in range(1, max_col + 1):
        col_letter = number_to_column_letter(col)
        header += f" | {col_letter}列"
    print(header)
    print("-" * 60)
    
    # 显示每行数据
    for row in range(1, max_row + 1):
        row_data = f"{row:3d}"
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is None:
                cell_str = "(空)"
            else:
                cell_str = str(cell_value)[:15]  # 只显示前15个字符
                if len(str(cell_value)) > 15:
                    cell_str += "..."
            row_data += f" | {cell_str:18s}"
        print(row_data)
    print("-" * 60)
    print(f"总行数：{sheet.max_row}，总列数：{sheet.max_column}")
    print()


def get_user_column_input(sheet, prompt_text, default_value=None):
    """
    获取用户输入的列号，并验证有效性
    
    参数：
        sheet: Excel工作表对象
        prompt_text: 提示信息
        default_value: 默认值（如果用户直接回车，使用此值）
    
    返回：
        有效的列号（数字）
    """
    max_col = sheet.max_column
    
    while True:
        if default_value is not None:
            user_input = input(f"{prompt_text}（直接回车使用默认值：{number_to_column_letter(default_value)}列）: ").strip()
            if user_input == "":
                return default_value
        else:
            user_input = input(f"{prompt_text}（输入列号，如A、B或1、2）: ").strip()
        
        # 转换为列号数字
        col_num = column_letter_to_number(user_input)
        
        if col_num is None:
            print("❌ 输入格式错误！请输入列号（如 A、B 或 1、2）")
            continue
        
        if col_num < 1 or col_num > max_col:
            print(f"❌ 列号超出范围！请输入 1 到 {max_col} 之间的列号（或 A 到 {number_to_column_letter(max_col)}）")
            continue
        
        return col_num


def translate_excel():
    """
    主函数：处理Excel文件，让用户选择翻译服务、源列和目标列，自动识别语言后互译
    中文会自动翻译成英文，英文会自动翻译成中文
    """
    try:
        # 让用户选择翻译服务
        print("=" * 60)
        print("请选择翻译服务：")
        print("  1. 有道翻译（Youdao）")
        print("  2. DeepL翻译")
        print("=" * 60)
        
        service_choice = input("请输入数字选择（1或2，直接回车默认使用有道翻译）: ").strip()
        
        if service_choice == '2':
            selected_service = 'deepl'
            service_name = 'DeepL'
            # 检查DeepL API密钥是否已配置
            if DEEPL_API_KEY == '你的DeepL_API_Key' or not DEEPL_API_KEY:
                print("❌ 错误：未配置DeepL API密钥！")
                print("请打开 translate_excel.py 文件，修改 DEEPL_API_KEY 配置")
                print("\n获取DeepL API密钥的方法：")
                print("  1. 访问 https://www.deepl.com/zh/pro-api")
                print("  2. 注册并登录账号")
                print("  3. 在账户中获取API密钥")
                return
        else:
            selected_service = 'youdao'
            service_name = '有道翻译'
            # 检查有道API密钥是否已配置
            if YOUDAO_APP_KEY == '你的AppKey' or YOUDAO_APP_SECRET == '你的AppSecret':
                print("❌ 错误：未配置有道翻译API密钥！")
                print("请打开 translate_excel.py 文件，修改 YOUDAO_APP_KEY 和 YOUDAO_APP_SECRET 配置")
                return
        
        print(f"✓ 已选择翻译服务：{service_name}\n")
        
        # 打开Excel文件
        print(f"正在打开Excel文件：{EXCEL_FILE}")
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        
        # 获取第一个工作表（sheet）
        sheet = workbook.active
        
        # 显示Excel文件预览，帮助用户了解文件结构
        show_excel_preview(sheet, max_cols=min(10, sheet.max_column))
        
        # 让用户选择要翻译的列（源列）
        source_column = get_user_column_input(
            sheet, 
            "📝 请输入要翻译的列号（源列）", 
            default_value=1  # 默认第一列
        )
        source_col_letter = number_to_column_letter(source_column)
        print(f"✓ 已选择源列：{source_col_letter}列（第{source_column}列）\n")
        
        # 让用户选择翻译结果填入的列（目标列）
        target_column = get_user_column_input(
            sheet,
            "📝 请输入翻译结果要填入的列号（目标列）",
            default_value=2  # 默认第二列
        )
        target_col_letter = number_to_column_letter(target_column)
        print(f"✓ 已选择目标列：{target_col_letter}列（第{target_column}列）\n")
        
        # 检查源列和目标列是否相同
        if source_column == target_column:
            print("⚠️  警告：源列和目标列相同，翻译结果会覆盖原文！")
            confirm = input("是否继续？（y/n）: ").strip().lower()
            if confirm != 'y' and confirm != 'yes':
                print("已取消操作")
                return
        
        # 询问是否从第一行开始（跳过标题行）
        print("\n是否跳过第一行（标题行）？")
        skip_header = input("输入 y 跳过第一行，直接回车从第一行开始翻译: ").strip().lower()
        start_row = 2 if skip_header in ['y', 'yes'] else 1
        
        if start_row == 2:
            print("✓ 将从第二行开始翻译（跳过标题行）")
        else:
            print("✓ 将从第一行开始翻译")
        
        # 询问是否调整翻译延时（用于避免频率限制）
        # 使用局部变量存储延时时间，避免修改全局变量
        current_delay = TRANSLATE_DELAY  # 使用全局变量作为默认值
        print(f"\n当前翻译延时设置为：{current_delay} 秒/次（固定1秒）")
        delay_input = input(f"是否调整延时时间？（直接回车使用默认值 {current_delay} 秒）: ").strip()
        
        if delay_input:
            try:
                custom_delay = float(delay_input)
                if custom_delay >= 0:
                    current_delay = custom_delay  # 使用局部变量
                    print(f"✓ 已设置延时时间为：{current_delay} 秒")
                else:
                    print(f"⚠ 延时时间不能为负数，使用默认值：{current_delay} 秒")
            except ValueError:
                print(f"⚠ 输入格式错误，使用默认值：{current_delay} 秒")
        else:
            print(f"✓ 使用默认延时时间：{current_delay} 秒")
        
        # 获取工作表中使用的最大行数
        max_row = sheet.max_row
        print(f"\n开始处理 {max_row - start_row + 1} 行数据...")
        print("=" * 60)
        
        # 遍历每一行
        success_count = 0  # 成功翻译的行数
        fail_count = 0  # 翻译失败的行数
        skip_count = 0  # 跳过的空行数
        
        for row_num in range(start_row, max_row + 1):
            # 获取源列的单元格值
            cell_value = sheet.cell(row=row_num, column=source_column).value
            
            # 检查单元格是否有内容
            if cell_value is None or str(cell_value).strip() == '':
                print(f"第 {row_num} 行 {source_col_letter}列为空，跳过")
                skip_count += 1
                continue
            
            # 将单元格值转换为字符串
            source_text = str(cell_value).strip()
            
            # 检查文本长度，如果过长则提前提示并跳过
            text_length = len(source_text)
            if text_length > 2000:
                print(f"第 {row_num} 行 ❌ 文本过长错误：文本长度 {text_length} 字符，超过2000字符限制")
                print(f"  跳过此行的翻译，建议手动缩短文本或分段处理")
                sheet.cell(row=row_num, column=target_column).value = f"文本过长错误（{text_length}字符，超过2000字符限制）"
                skip_count += 1
                continue
            
            # 自动检测文本语言（中文还是英文）
            detected_lang = detect_language(source_text)
            
            # 根据检测到的语言确定翻译方向（使用统一的语言代码格式）
            if detected_lang == 'zh':
                # 如果是中文，翻译成英文
                from_lang_code = 'zh'
                to_lang_code = 'en'
                lang_info = "中文 → 英文"
            elif detected_lang == 'en':
                # 如果是英文，翻译成中文
                from_lang_code = 'en'
                to_lang_code = 'zh'
                lang_info = "英文 → 中文"
            else:
                # 如果无法判断语言，默认按中文处理
                from_lang_code = 'zh'
                to_lang_code = 'en'
                lang_info = "未知语言，默认：中文 → 英文"
                print(f"  ⚠ 无法判断第 {row_num} 行的语言类型，将按中文处理")
            
            # 显示当前处理的行和翻译方向，同时显示文本长度
            text_length = len(source_text)
            text_preview = source_text[:30] + "..." if len(source_text) > 30 else source_text
            print(f"正在翻译第 {row_num} 行 [{lang_info}]（文本长度：{text_length}字符）：{text_preview}")
            
            # 调用统一的翻译函数，传入检测到的语言方向和选择的翻译服务
            translated_text = translate_text(source_text, from_lang_code, to_lang_code, selected_service)
            
            if translated_text:
                # 如果翻译成功，将结果写入目标列
                sheet.cell(row=row_num, column=target_column).value = translated_text
                print(f"  ✓ 翻译成功：{translated_text}")
                success_count += 1
            else:
                # 如果翻译失败，在目标列写入提示信息（或者留空）
                sheet.cell(row=row_num, column=target_column).value = "翻译失败"
                print(f"  ✗ 翻译失败")
                fail_count += 1
                # 翻译失败后，等待1秒再继续下一行
                if row_num < max_row:
                    print(f"  ⏸ 翻译失败，等待 1 秒后继续下一行...")
                    time.sleep(1.0)  # 失败后等待1秒
            
            # 添加延时，避免API调用过于频繁（有道API有频率限制）
            # 使用用户设置的延时时间（current_delay）
            if row_num < max_row:  # 最后一行不需要延时
                time.sleep(current_delay)
        
        # 保存修改后的Excel文件
        print("\n" + "=" * 60)
        print(f"正在保存文件...")
        workbook.save(EXCEL_FILE)
        print(f"✓ 文件已保存！")
        print(f"\n📊 统计信息：")
        print(f"  成功翻译：{success_count} 行")
        print(f"  翻译失败：{fail_count} 行")
        print(f"  跳过空行：{skip_count} 行")
        print(f"  总计处理：{success_count + fail_count + skip_count} 行")
        
    except FileNotFoundError:
        print(f"❌ 错误：找不到文件 '{EXCEL_FILE}'，请检查文件路径是否正确")
    except Exception as e:
        print(f"❌ 处理Excel文件时出现错误：{str(e)}")


if __name__ == '__main__':
    """
    程序入口：运行主函数（API密钥检查在主函数中根据选择的翻译服务进行）
    """
    print("=" * 60)
    print("Excel 中英互译工具")
    print("功能：自动识别中文/英文，然后互译")
    print("支持：有道翻译 / DeepL翻译")
    print("=" * 60)
    translate_excel()
    print("=" * 60)
    print("程序执行完毕！")
    print("=" * 60)

